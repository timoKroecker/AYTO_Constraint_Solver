import openpyxl

# [first row/col, last row/col + 1, step_size]
WB_NIGHTS = [19, 29, 1]

RED = "FF0000"
GREEN = "00A933"
YELLOW = "FFEE78"
WHITE = "000000"

def load_worksheet(path, sheet_name, data_only=False):
    workbook = openpyxl.load_workbook(path, data_only=data_only)
    return workbook, workbook[sheet_name]

def get_girl_index(girl):
    return str(girl + 2)

def get_boy_index(boy):
    return chr(boy + 98)

def get_cell_index(girl, boy):
    return chr(boy + 98) + str(girl + 2)

def get_name_dictionaries(sheet):
    girls = {}
    boys = {}
    max_num = 15
    for i in range(max_num):
        index = get_girl_index(i)
        name = sheet["a" + str(index)].value
        if(name == None or name == ""):
            break
        girls[name.replace(" ", "")] = i + 1
    for i in range(max_num):
        index = get_boy_index(i)
        name = sheet[index + "1"].value
        if(name == None or name == ""):
            break
        boys[name.replace(" ", "")] = i + 1
    return girls, boys

def load_input(path):
    map_ = {
    RED: -1,
    GREEN: 1,
    WHITE: 0
    }
    _, sheet = load_worksheet(path, "Input", data_only=True)
    girls, boys = get_name_dictionaries(sheet)
    input_matrix = []
    for i in range(len(girls)):
        row = []
        for j in range(len(boys)):
            cell_index = get_cell_index(i, j)
            hex_color = sheet[cell_index].fill.fgColor.index[2:]
            row.append(map_[hex_color])
        input_matrix.append(row)

    matching_nights = []
    lights = []
    num_couples = min(len(girls), len(boys))
    lights_col = chr((ord("b") + 2 * num_couples))
    for night in range(WB_NIGHTS[0], WB_NIGHTS[1], WB_NIGHTS[2]):
        light = sheet[lights_col + str(night)].value
        if(light == None):
            break
        lights.append(int(light))
        row = []
        for i in range(98, 97 + 2 * num_couples, 2):
            girl_name = sheet[chr(i + 1) + str(night)].value.replace(" ", "")
            boy_name = sheet[chr(i) + str(night)].value.replace(" ", "")
            row.append([girls[girl_name], boys[boy_name]])
        matching_nights.append(row)

    third_wheel = sheet["b15"].value
    if(third_wheel is None):
        third_wheel = -1
    else:
        third_wheel = girls[third_wheel]

    return input_matrix, matching_nights, lights, third_wheel

def hex_to_rgb(hex_color):
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

def linear_map(value, domain, range_):
    alpha = (value - domain[0]) / (domain[1] - domain[0])
    return range_[0] + alpha * (range_[1] - range_[0])

def diverging_map(value, domain, range_start, range_mid, range_end):
    mid_domain = (domain[1] - domain[0]) / 2
    if(value < mid_domain):
        return linear_map(value, [domain[0], mid_domain], [range_start, range_mid])
    return linear_map(value, [mid_domain, domain[1]], [range_mid, range_end])

def get_hex_color(probability):
    color = [0, 0, 0]
    red = hex_to_rgb(RED)
    yellow = hex_to_rgb(YELLOW)
    green = hex_to_rgb(GREEN)
    for i in range(len(color)):
        color[i] = int(diverging_map(probability, [0, 1], red[i], yellow[i], green[i]))
    return '%02x%02x%02x' % tuple(color)

def write_probabilities(path, probabilities):
    _, read_sheet = load_worksheet(path, "Output", data_only=True)
    write_book, write_sheet = load_worksheet(path, "Output")
    for i in range(len(probabilities)):
        for j in range(len(probabilities[0])):
            cell_index = get_cell_index(i, j)
            probability = probabilities[i, j]
            write_sheet[cell_index].value = str(round(probability * 100, 2)) + "%"

            hex_color = get_hex_color(probability)
            fill_color = openpyxl.styles.colors.Color(rgb='00' + hex_color)
            fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=fill_color)
            write_sheet[cell_index].fill = fill

    girls, boys = get_name_dictionaries(read_sheet)
    num_couples = min(len(girls), len(boys))
    for night in range(WB_NIGHTS[0], WB_NIGHTS[1], WB_NIGHTS[2]):
        for i in range(98, 97 + 2 * num_couples, 2):
            girl_index = chr(i + 1) + str(night)
            boy_index = chr(i) + str(night)
            girl_name = read_sheet[girl_index].value
            boy_name = read_sheet[boy_index].value
            if(girl_name == None or boy_name == None):
                continue
            girl_name = girl_name.replace(" ", "")
            boy_name = boy_name.replace(" ", "")
            cell_index = get_cell_index(girls[girl_name] - 1, boys[boy_name] - 1)
            hex_color = write_sheet[cell_index].fill.fgColor.index[2:]
            fill_color = openpyxl.styles.colors.Color(rgb='00' + hex_color)
            fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=fill_color)
            write_sheet[girl_index].fill = fill
            write_sheet[boy_index].fill = fill
    
    write_book.save(path)